from multiprocessing import Pool, freeze_support, Manager
import pdfplumber
import re
import os
import time
import datetime
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import traceback
import multiprocessing
from functools import partial
from threading import Thread
from concurrent.futures import ThreadPoolExecutor


class PdfExtracter:
    def __init__(self):
        m = Manager()
        self.queue = m.Queue()
        self.queueErr = m.Queue()
        self.errorsFls = m.Queue()
        self.lock = m.Lock()
        self.counter = m.Value('i', 0)
        self.max = None
        
    def set_tracking_text(self, pdftext):
        mtch = re.search(r'[ ]*?(\d{10,20})[A-Z0-9 -,]+U[ ]*?S[ ]*?', pdftext)
        if mtch is not None:
            index = mtch.start()
            return index - 5
        return 0

    def aryse_regex(self, pattern, txt, end):
        matched = re.search('{}'.format(pattern), txt)
        try:
            end = end + matched.end()
            return matched.group(1).strip(), end
        except:
            return None, end
    
    def get_address_aryse(self, txt):
        addr = txt.split('\n')
        res = ''
        # for i, adr in enumerate(addr):
        #     if adr.startswith('Order Number') or adr.startswith('Tracking'):
        #         pass		
        #     elif adr.startswith('Ship To'): 
        #         if not adr.strip().__contains__('Order Number'):
        #             res += adr[8:-5].strip()
        #     else:
        #         if adr.strip().__contains__('Tracking Number'):
        #             res += adr.split('Tracking Number')[0].strip()
        #         elif i == 0:
        #             res += adr.strip()
        #         else:
        #             res += ' '.join(adr.split(' ')[:-1])
        # return res
        # res1 = ''
        for adr in addr:
            if adr.strip().startswith('Ship To') or adr.strip().startswith('Tracking Number'):
                pass
            else:
                if adr.strip().__contains__('Tracking Number'):
                    res1 += adr.split('Tracking Number')[0].strip()
                else:
                    res1 += adr.strip()
        return res
        # print(res, res1)
        
    def get_data_aryse(self, pdf):
        try:
            with pdfplumber.open(pdf) as pdff:
                for pg in pdff.pages:
                    try:
                        txt = pg.extract_text()
                        try:
                            tab = pg.extract_tables()[0]
                        except:
                            if txt is None:  # no text, page contain image instead
                                self.errorsFls.put(str(pdf))  # f'File: {pdf}  - Contain Images\n-------------------------------------------------\n'
                                return # got cid format
                            self.errorsFls.put(str(pdf))  # f'File: {pdf}  - CID Format error\n-------------------------------------------------\n'
                            return # got cid format
                        if txt.strip().startswith('S T A T E M E N T'):
                            self.errorsFls.put(str(pdf))  # f'File: {pdf}  - HC Statement\n-------------------------------------------------\n'
                            return
                        patient_name, end = self.aryse_regex(r'Patient *?\n*?[Name:]*?([A-Za-z \-,\.\?]+)Patient', txt, 0)  # Patient\s?Name:([A-Za-z -,]+)Patient
                        order_date, endod = self.aryse_regex(r'Order Date: *?\n*?[A-Za-z0-9 \-,\.\?]*?([A-Za-z]+ *?\d{1,2},? *?\d{4})', txt[end-1:], end)  # Order Date: *?\n*?([A-Za-z ]+\d{1,2},? *?\d{4})
                        order_number, end = self.aryse_regex(r'Order Number: *?\n*?[A-Za-z0-9 :]*([0-9 ]+)', txt[endod-1:], endod)  # Order Number:([0-9 ]+)
                        tracking_number, end = self.aryse_regex(r'Tracking ?[Number:]*\n*? *?([A-Za-z0-9 ,-\?]+)', txt[end-1:], end)  # Tracking Number:([A-Za-z0-9 ]+)
                        address = self.get_address_aryse(txt[endod:end])
                        tab = tab[1:]
                        price = len(tab) * 400
                        tracking_number = tracking_number.split(' ')[-1]
                        self.queue.put((patient_name, address, order_date, order_number, tracking_number, '${:,.2f}'.format(price), tab))
                    except:
                        err = traceback.format_exc()
                        self.queueErr.put(f'File: {pdf}\nFullname : {str(patient_name)}\nException : 1\n{err}-------------------------------------------------\n')
        except Exception as e:
            err = traceback.format_exc()
            self.queueErr.put(f'File: {pdf}\nException : 2\n{err}-------------------------------------------------\n')
        finally:
            self.lock.acquire()
            self.counter.value += 1
            self.lock.release()

    def get_data_usd_invoice(self, pdf):
        pdftext = ''
        VERTALOC_INV = ''
        with pdfplumber.open(pdf) as pdf:
            table = []
            last = len(pdf.pages)
            for ind, pg in enumerate(pdf.pages, start=1):
                if ind == 1:
                    txt = pg.extract_text()
                    date, invoice = re.findall(r'\d+/\d+/\d+ \d+', txt)[0].split()
                    VERTALOC_INV = re.findall(r'VERTALOC INV[0-9.# ]+', txt)[0].replace('VERTALOC INV.', '').replace('#', '').replace(' ', '')
                des = pg.extract_tables()[2][1][2]
                pdftext = pdftext + des + '\n'
            Patient_list_page = re.finditer(r'Patient Full Name:(.+)', pdftext)
            adress_list_page = re.finditer(r"Patient Full Address:(.+\n*?.*?)\n*?Patient*?", pdftext)
            brace_list_page = re.finditer(r'Patient Brace Line \d:(.+)', pdftext)
            index_tracking = self.set_tracking_text(pdftext)
            tracking = pdftext[index_tracking:].replace(' ', '').strip()
        return pdftext, tracking, date, invoice, VERTALOC_INV, Patient_list_page, adress_list_page, brace_list_page

    def get_data_invoice(self, pdf):
        pdftext = ''
        with pdfplumber.open(pdf) as pdff:
            last = len(pdff.pages)
            for ind, pg in enumerate(pdff.pages, start=1):
                if ind == 1:
                    txt = pg.extract_text()
                    date, invoice = re.findall(r'\d+/\d+/\d+ \d+', txt)[0].split()
                des = pg.extract_tables()[1][3:]
                for d in des:
                    if d[2] is None:
                        break
                    else:
                        pdftext = pdftext + d[2] + '\n'
            Patient_list_page = re.finditer(r'Patient Full Name:(.+)', pdftext)
            adress_list_page = re.finditer(r"Patient Full Address:(.+\n*?.*?)\n*?Patient*?", pdftext)
            brace_list_page = re.finditer(r'Patient Brace Line \d:(.+)', pdftext)
            index_tracking = self.set_tracking_text(pdftext)
            tracking = pdftext[index_tracking:].replace(' ', '').strip()
        return pdftext, tracking, date, invoice, Patient_list_page, adress_list_page, brace_list_page


    def get_data(self, pdf, company):
        try:
            VERTALOC_INV = ''
            errormsg = ''
            if company == 'USD_INVOICE':
                pdftext, tracking, date, invoice, VERTALOC_INV, Patient_list_page, adress_list_page, brace_list_page = self.get_data_usd_invoice(pdf)
            else:
                pdftext, tracking, date, invoice, Patient_list_page, adress_list_page, brace_list_page = self.get_data_invoice(pdf)

            for names, adress, brace in zip(Patient_list_page, adress_list_page, brace_list_page):
                try:
                    names = names.group(1)
                    # WILL ADD SPLIT(' ', 1)
                    fullname = names.replace('"', '').replace(' ,', ',').strip()
                    nmsplt = fullname.replace(',', '').split(' ')
                    first = nmsplt[0].strip().upper()
                    if len(nmsplt) > 2:
                        last =  ''.join(set(list((nmsplt[1] + nmsplt[2]).replace(' ', '').strip()))).upper()
                    else:
                        last =  ''.join(set(list((nmsplt[1]).replace(' ', '').strip()))).upper()
                    patient = f'[0-9 ]+{first}[' + last + ' -]+'
                    track_str = re.findall(r'{}'.format(patient), tracking)[0].split(' ')[0]
                    track = re.findall(r'\d+', track_str)[0]
                    self.queue.put((invoice.strip(), date.strip(), VERTALOC_INV.strip(), fullname.strip(), adress.group(1).replace('\n', ' ').replace('"', ' ').strip(), brace.group(1).strip(), '${:,.2f}'.format(400.0), track.strip()))
                except:
                    err = traceback.format_exc()
                    errormsg += f'File: {pdf}\nFullname : {fullname}\nTracking regex : {patient}\nException : 1\n{err}-------------------------------------------------\n'
                    self.queueErr.put(errormsg)
                
        except Exception as e:
            err = traceback.format_exc()
            errormsg += f'File: {pdf}\nFullname : {fullname}\nTracking regex : {patient}\nException : 2\n{err}-------------------------------------------------\n'
            self.queueErr.put(errormsg)
        finally:
            self.lock.acquire()
            self.counter.value += 1
            self.lock.release()
        
    def set_sheet_title(self, mySheet, ttl, A, B):
        myFont = Font(name='Calibri',
                                size=20,
                                bold=True,
                                italic=True,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')
        alignment = Alignment(horizontal='center',
                                  vertical='center',
                                  text_rotation=0,
                                  wrap_text=True,
                                  shrink_to_fit=False,
                                  indent=0)
        mySheet[f'{A}1'].font = myFont
        mySheet[f'{A}1'].alignment = alignment
        mySheet[f'{A}1'] = ttl
        mySheet.merge_cells(f'{A}1:{B}1')
        week_yr = datetime.datetime.today().isocalendar()[1]
        mySheet[f'{A}2'].font = myFont.size = 16
        mySheet[f'{A}2'].alignment = alignment
        mySheet[f'{A}2'] = f'CURRENT WEEK :  {week_yr}'
        mySheet.merge_cells(f'{A}2:{B}2')
    
    def write2Excel(self, name, inv, table, company=None):
        alignment = Alignment(horizontal='center',
                                  vertical='top',
                                  text_rotation=0,
                                  wrap_text=True,
                                  shrink_to_fit=False,
                                  indent=0)
        data = []
        wb = Workbook()
        wb.active.title = inv
        sheet = wb.active
        sheet = self.excel_Formating(sheet, company)
        count = 4
        while not table.empty():
            r = list(table.get())
            pbq = r.pop()
            length = len(pbq)
            strt = None
            for ind, (p, b, q) in enumerate(pbq):
                if ind == 0:
                    r.extend((p, b, q))
                    sheet.append(r)
                    count += 1
                    strt = count
                else:
                    sheet.append((None,None,None,None,None, None, p, b, q))
                    count += 1
                for alp in string.ascii_uppercase:
                    if company == 'ARYSE':
                        if alp == 'J':
                            break
                        elif ind > 0 and ind == (length - 1):
                            if alp not in ['G', 'H', 'I']:
                                sheet.merge_cells(f'{alp}{strt}:{alp}{count}')
                    else:
                        if alp == 'I':
                            break
                    sheet[f'{alp}{count}'].alignment = alignment
                    
        sheet.freeze_panes = sheet['A5']
        max_row = sheet.max_row
        if company == 'ARYSE':
            ref = f"A4:I{max_row}"
        else:
            ref = f"A4:H{max_row}"
        sheet.auto_filter.ref = ref
        wb.save(name)
        
    def excel_Formating(self, sheet, company):
        alignment = Alignment(horizontal='center',
                                  vertical='top',
                                  text_rotation=0,
                                  wrap_text=True,
                                  shrink_to_fit=False,
                                  indent=0)
        alignment2 = Alignment(horizontal='center',
                                  vertical='center',
                                  text_rotation=180,
                                  wrap_text=True,
                                  shrink_to_fit=False,
                                  indent=0)
        bold11Font = Font(name='Calibri',
                                size=11,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')
        fill = PatternFill(start_color='cccccc', end_color='cccccc', fill_type = "mediumGray")
        border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
        if company == 'ARYSE':
            header = ('Patient Name','Adrress','Date','Order No','Tracking No', 'PRICE', 'Product', 'Brace', 'Quantity')
            widths = [20, 50, 18, 10, 20, 15, 20, 40, 5]
        else:
            header = ('INVOICE NO', 'DATE', 'VERTALOC INV', 'PATIENT NAME', 'PATIENT ADRRESS', 'BRACE', 'PRICE', 'TRACKING NO')
            widths = [10, 15, 15, 20, 50, 40, 15, 15]
        ct = time.strftime("%b %d %Y")
        sheet.row_dimensions[1].height = 15
        sheet.row_dimensions[2].height = 15
        self.set_sheet_title(sheet, f'{company} Reports  - {ct}', 'C', 'E')
        for col, alp, wid  in zip(header, string.ascii_uppercase, widths):
            sheet.column_dimensions[alp].width = wid
            sheet.row_dimensions[4].height = 60
            sheet[f'{alp}4'] = col
            sheet[f'{alp}4'].font = bold11Font
            sheet[f'{alp}4'].fill = fill
            sheet[f'{alp}4'].border = border
            if company == 'ARYSE' and alp == 'I':
                sheet[f'{alp}4'].alignment = alignment2
            else:
                sheet[f'{alp}4'].alignment = alignment
        return sheet
        
        
    def write_errors(self, queue):
        if not queue.empty():
            with open('errors.txt', "w") as f:
                while not queue.empty():
                    f.write(queue.get())
    
    def move_Errors_Files(self, queue):
        while not queue.empty():
            file_path = queue.get()
            os.rename(file_path, os.path.join('PdfIssues', os.path.basename(file_path)))
        
    def multi_processing(self, fls, company):
        freeze_support()
        with ThreadPoolExecutor() as thrds:  # multiprocessing.pool.ThreadPool() #concurrent.futures.ThreadPoolExecutor()
            if company == 'ARYSE':
                thrds.map(self.get_data_aryse, fls)
            else:
                prc = partial(self.get_data, company=company) # , queue=queue, queueErr=queueErr, lock=lock, counter=counter
                thrds.map(prc, fls)
            

    def multitasking_manager(self, files, company):
        cpu_count = multiprocessing.cpu_count() - 1
        chrnk_size = int(len(files) / cpu_count)
        try:
            fils = [files[i:i + chrnk_size] for i in range(0, len(files), chrnk_size)]
        except ValueError:
            fils = [files]
        if len(files) <= cpu_count:
            cpu_count = len(files)
        with Pool(cpu_count) as p:
            parcial = partial(self.multi_processing, company=company) # , queue=self.queue, queueErr=self.queueErr, lock=self.lock, counter=self.counter
            p.map(parcial, fils)
        self.write_errors(self.queueErr)
        ct = time.strftime("-%Y%m%d-%H%M%S")
        self.write2Excel(company + ct + '.xlsx', company, self.queue, company=company)
        self.move_Errors_Files(self.errorsFls)
        


# for i, (product, brace, quantity) in enumerate(tab):
    # if i == 0:
    #     pass
    # elif i == 1:
    #     queue.put((patient_name, address, order_date, order_number, tracking_number, product, brace, quantity))
    # else:
    #     queue.put(('','','','','',product, brace, quantity))